from openpyxl import Workbook
import time


"""
book = Workbook()
sheet = book.active

sheet['A1'] = 56
sheet['A2'] = 43

now = time.strftime("%x")
sheet['A3'] = "=A1+A2"

book.save("sample.xlsx")


"""
"""
from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)

book.save('appending.xlsx')


"""
"""

from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

data = [
    ['Item', 'Colour'],
    ['pen', 'brown'],
    ['book', 'black'],
    ['plate', 'white'],
    ['chair', 'brown'],
    ['coin', 'gold'],
    ['bed', 'brown'],
    ['notebook', 'white'],
]

for r in data:
    sheet.append(r)

sheet.auto_filter.ref = 'A1:B8'
sheet.auto_filter.add_filter_column(1, ['brown', 'white'])
sheet.auto_filter.add_sort_condition('B2:B8')

wb.save('filtered.xlsx')

"""


"""

import openpyxl

book = openpyxl.load_workbook('sheets.xlsx')

print(book.get_sheet_names())

active_sheet = book.active
print(type(active_sheet))

sheet = book.get_sheet_by_name("March")
print(sheet.title)
"""



"""
from openpyxl import Workbook
from openpyxl.styles import Alignment

book = Workbook()
print("Book: ", book)
sheet = book.active

sheet.freeze_panes = 'B2'

print("sheet: ", sheet)

book.save('freezing.xlsx')
"""


a= 100
b= 151

c= a // b

print("c : ", c)